#encoding=utf-8

import argparse
import pkuseg
import pprint

import sys
import os
# 确保在命令行下执行，不出现ModuleNotFoundError: No module named ERROR
sys.path.append(os.getcwd())    # 在命令行中，把当前路径加入到sys.path中
print(sys.path)
from openpyxl import load_workbook


def get_label_maps(path):
    wb = load_workbook(path)
    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[0])

    id2name = {}        # 标签Id 到 标签名字  字典表
    id_tree = {}        # 标签树结构

    # 把数据存到字典中
    last_id = 0
    for rx in range(1, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w2 = ws.cell(row=rx, column=2).value
        w3 = ws.cell(row=rx, column=3).value
        w4 = ws.cell(row=rx, column=4).value
        if w1 == "标签ID":
            continue
        if w1 is not None:
            id2name[w1] = w2
            id_tree[w1] = []
            last_id = w1
        if w3 is not None:
            id2name[w3] = w4

        id_tree[last_id].append(w3)

    print("label2name dict:")
    print(id2name)
    print(id_tree)
    return id2name, id_tree


dict_file = 'data/my_dict.txt'

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.register("type", "bool", lambda v: v.lower() == "true")
    parser.add_argument(
        "--is_debug",
        type=lambda x: (str(x).lower() == 'true'),
        default=True,
        help="use debug mode")

    FLAGS, unparsed = parser.parse_known_args()
    pp = pprint.PrettyPrinter().pprint
    pp("FLAGS: " + str(FLAGS))
    pp("unparsed: " + str(unparsed))
    print(FLAGS.is_debug)

    finnal_context = []

    num = 0
    # raw_tribe_data.txt文件在160上，太大
    input_file = "data/test.txt" if FLAGS.is_debug else "data/raw_tribe_data.txt"
    with open(input_file, "r", encoding='utf-8') as f:
        for line in f:
            sp = line.split("\t")
            label = sp[0].replace("__label__", "")
            if label is not None and label != "" and label.isdigit() and int(label) > 800 and int(label) <= 1000:
                context = []
                for i in sp:
                    if i.startswith("__label__"):
                        continue
                    if i.startswith("__infoid__") or i.startswith("__tribeid__") or i.startswith("__topic__"):
                        break
                    without_answer = " ".join(i.split("\t")[0:2])
                    context.append(without_answer)

                num += 1
                if num % 100 == 0:
                    print("processing {} lines".format(num))

                context_str = " ".join(context)
                finnal_context.append((label, context_str))

    print("valid file lines total num: {}".format(len(finnal_context)))

    id2name, id_tree = get_label_maps("data/new_labels.xlsx")

    import pandas as pd
    df = pd.DataFrame(finnal_context, columns=["label", "context"])
    print("-" * 20)
    print(df.sample(5))

    df["label"] = df["label"].apply(lambda x: id2name[int(x)])

    print("-" * 10 + "label -> num" + "-" * 10)
    for label_df in df.groupby("label"):
        print("label: {} nums: {}".format(label_df[0], len(label_df[1])))
    print("-" * 20)

    input_file = 'input.txt'
    output_file = 'output.txt'
    final_file = 'old_finnal_data.txt'

    df.to_csv(input_file, sep=" ", index=False, header=False, quotechar=" ")

    print("seg txt start:")
    pkuseg.test(input_file, output_file, user_dict=dict_file, nthread=20)
    print("seg txt stop")
    print("-" * 20)

    with open(final_file, 'w') as out_f:
        with open(output_file, 'r') as input_f:
            for line in input_f:
                sp = line.split(" ", 1)
                if len(sp) >= 2:
                    out_f.write("__label__" + sp[0] + "\t" + sp[1] + "\n")

    print("final file:{}".format(final_file))

